import time
import traceback

from django.core.paginator import Paginator
from django.db.models import Q
from django.http import HttpResponse,JsonResponse,FileResponse
from django.utils.encoding import escape_uri_path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font,Border,Side,Alignment

from .models import basedata
from openpyxl.compat import range
from .views import token_decorator
from django.urls import resolve
from lxml.doctestcompare import strip
import requests,urllib3,openpyxl,pymysql,csv,datetime,codecs,json

def import_user(self, request, obj, change):
    wb = load_workbook(filename=obj.file)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['productcode', 'name', 'model', 'description','RRP_USD','RRP_EUR','RRP_GBP','RRP_RMB']
    lists = []
    rows=ws.max_row
    for row in range(2, rows+1):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    sqllist = []
    for cell in lists:
        # for header in headers:
        productcode = cell['productcode']
        name = cell['name']
        model = cell['model']
        description = cell['description']
        RRP_USD = cell['RRP_USD']
        RRP_EUR = cell['RRP_EUR']
        RRP_GBP = cell['RRP_GBP']
        RRP_RMB = cell['RRP_RMB']
        sql = basedata(productcode=productcode, name=name, model=model, description=description, RRP_USD=RRP_USD,RRP_EUR=RRP_EUR,RRP_GBP=RRP_GBP,RRP_RMB=RRP_RMB)
        sqllist.append(sql)
    basedata.objects.bulk_create(sqllist)

def download_user(request):
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    sql = 'select productcode,name,model,description,product_type,price_source,RRP_USD,RRP_EUR,RRP_GBP,RRP_RMB,createddate,lastmodifieddate,remark from myapp_basedata '
    cur.execute(sql)
    data = list(cur.fetchall())
    response = list_to_excel(data,lang='all')
    return response

def update_user(self,request,obj,change):
    wb = load_workbook(filename=obj.update_file)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['productcode', 'name', 'model', 'description', 'RRP_USD', 'RRP_EUR', 'RRP_GBP', 'RRP_RMB','createddate','lastmodifieddate']
    lists = []
    rows = ws.max_row
    for row in range(2, rows + 1):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    for cell in lists:
        # for header in headers:
        productcode = cell['productcode']
        name = cell['name']
        model = cell['model']
        description = cell['description']
        RRP_USD = cell['RRP_USD']
        RRP_EUR = cell['RRP_EUR']
        RRP_GBP = cell['RRP_GBP']
        RRP_RMB = cell['RRP_RMB']
        lastmodifieddate = datetime.datetime.now()
        basedata.objects.filter(productcode=productcode).update(name=name, model=model, description=description, RRP_USD=RRP_USD,
                       RRP_EUR=RRP_EUR, RRP_GBP=RRP_GBP, RRP_RMB=RRP_RMB,lastmodifieddate=lastmodifieddate)

@token_decorator
def upsert_user(request):
    res = {'code':20000, 'msg':'success'}
    try:
        uploadedFile = request.FILES.get('file', None)
        print(uploadedFile)
        wb = load_workbook(filename=uploadedFile)
        ws = wb.get_sheet_names()
        ws = wb.get_sheet_by_name(ws[0])
        rows = ws.max_row
        cols = ws.max_column
        headers = ['productcode', 'name', 'model', 'description', 'product_type', 'price_source', 'RRP_USD', 'RRP_EUR',
                   'RRP_GBP', 'RRP_RMB', 'remark']
        lists = []
        if rows - 1 == 0 and cols - 1 == 0:
            pass
        else:
            sqllist = []
            for i in range(2, rows + 1):
                r = {}
                for j in range(1, len(headers) + 1):
                    key = headers[j - 1]
                    cellvalue = ws.cell(i, j).value
                    if cellvalue !=None:
                        r[key] = str(cellvalue)
                    else:
                        r[key] = ''
                lists.append(r)
            db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
            cur = db.cursor()
            for cell in lists:
                cell['lastmodifieddate'] = datetime.datetime.now()
                sqllist = sqllist + upsert_basedata(cell)
            if len(sqllist) > 0:
                basedata.objects.bulk_create(sqllist)
            res['msg'] = upsert_cpq_items(lists)
    except Exception as e:
        print(e)
        print('llllllllllllllllll',traceback.print_exc())
        res['code'] = 2005
        res['msg'] = '推送至CPQ失败！'
    return JsonResponse(res)

@token_decorator
def searchall(request):
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    results = []
    jsonData = []
    if request.method == "GET":
        start = request.GET.get("item", None)
        page_size = request.GET.get('page_size', 10)
        page_num = request.GET.get('page_num', 1)
        if start is not None and len(str(start)) != 0:
            starts = start.split(',')
            for i in starts:
                starts = strip(i)
                print(starts)
                sql = "SELECT id,productcode,name,model,description,product_type,price_source,RRP_USD,RRP_EUR,RRP_GBP,RRP_RMB,createddate,lastmodifieddate,remark FROM myapp_basedata WHERE productcode = '" \
                      + starts + "' OR name like '%" + starts + "%'" + " OR model like '%" + starts + "%'"
                try:
                    cur.execute(sql)
                    results = results + list(cur.fetchall())
                    print(results)
                except Exception as e:
                    print(traceback.print_exc())
                    raise e
        else:
            sql = 'SELECT id,productcode,name,model,description,product_type,price_source,RRP_USD,RRP_EUR,RRP_GBP,RRP_RMB,createddate,lastmodifieddate,remark from myapp_basedata'
            cur.execute(sql)
            results = cur.fetchall()
        total = len(results)
        paginator = Paginator(results, page_size)
        context = []
        try:
            if paginator.page(page_num):
                page_of_price = paginator.page(page_num)
                context = list(page_of_price)
        except Exception as e:
            print(e)
            print(traceback.print_exc())
            page_num = 1
            page_of_price = paginator.page(page_num)
            context = list(page_of_price)
        # print(context)
        for row in context:
            res = {}
            res['id'] = row[0]
            res['productcode'] = row[1]
            res['name'] = row[2]
            res['model'] = row[3]
            res['description'] = row[4]
            res['product_type'] = row[5]
            if row[6] == 'Pricing':
                res['price_source'] = '定价'
            elif row[6] == 'Quotation':
                res['price_source'] = '报价'
            else:
                res['price_source'] = row[6]
            res['RRP_USD'] = row[7]
            res['RRP_EUR'] = row[8]
            res['RRP_GBP'] = row[9]
            res['RRP_RMB'] = row[10]
            res['createddate'] = row[11].strftime('%Y-%m-%d %H:%M:%S')
            res['lastmodifieddate'] = row[12].strftime('%Y-%m-%d %H:%M:%S')
            res['remark'] = row[13]
            jsonData.append(res)
        # print(jsonData)
        return JsonResponse({'rows': jsonData,'page_num':int(page_num),'total':total})
    else:
        return JsonResponse({'rows': jsonData})

@token_decorator
def price_list(request):
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    sql = 'select productcode,name,model,description,product_type,price_source,RRP_USD,RRP_EUR,RRP_GBP,RRP_RMB,createddate,lastmodifieddate,remark,id from myapp_basedata'
    cur.execute(sql)
    data = cur.fetchall()
    total = len(data)
    page_size = request.GET.get('page_size',10)
    paginator = Paginator(data,page_size)
    page_num = request.GET.get('page_num',1)
    context = []
    try:
        if paginator.page(page_num):
            page_of_price = paginator.page(page_num)
            context=list(page_of_price)
    except Exception as e:
        print(e)
        page_num = 1
        page_of_price = paginator.page(page_num)
        context = list(page_of_price)
    # print(context)
    jsonData = []
    for row in context:
        res = {}
        res['productcode'] = row[0]
        res['name'] = row[1]
        res['model'] = row[2]
        res['description'] = row[3]
        res['product_type'] = row[4]
        if row[5] == 'Pricing':
            res['price_source'] = '定价'
        elif row[5] == 'Quotation':
            res['price_source'] = '报价'
        else:
            res['price_source'] = row[5]
        res['RRP_USD'] = row[6]
        res['RRP_EUR'] = row[7]
        res['RRP_GBP'] = row[8]
        res['RRP_RMB'] = row[9]
        res['createddate'] = row[10].strftime('%Y-%m-%d %H:%M:%S')
        res['lastmodifieddate'] = row[11].strftime('%Y-%m-%d %H:%M:%S')
        res['remark'] = row[12]
        res['id'] = row[13]
        jsonData.append(res)
    # print(jsonData)
    return JsonResponse({'rows':jsonData,'page_num':int(page_num),'total':total})

item_list = []
@token_decorator
def bom_list(request):
    global item_list
    try:
        bomId = request.GET.get('bomId',None)
        token = request.GET.get('token',None)
        lang = request.GET.get('lang', None)
        language = {'language':lang}
        print(bomId,token,lang)
        url = 'https://xxxxxxxxxxxxxx.com/hvpc-mdata/v1/11/3/hvpc/mdm/bom/path/tree-bom-val'
        url2 = 'https://xxxxxxxxxxxxxx.com/iam/hzero/v1/users/default-language'
        params = {'bomId':bomId}
        Authorization = request.META.get('HTTP_AUTHORIZATION')
        print(Authorization)
        headers = {
            'Content-Type':'application/x-www-form-urlencoded;charset=UTF-8',
            'Authorization':token
        }
        urllib3.disable_warnings()
        requests.put(url2, data=language, headers=headers, verify=False)
        result = requests.get(url,params=params,headers=headers,verify=False)
        bomlist = result.json()[0]
        a = []
        # a.append(bomlist['path'])
        a.append(bomlist['itemCode'])
        a.append(bomlist['itemName'])
        a.append(bomlist['itemDescription'])
        valueFlag = 'valueFlag'
        if valueFlag in bomlist.keys():
            a.append(bomlist['valueFlag'])
        else:
            a.append('0')
        item_list.append(a)
        get_itemcode(bomlist)
        # print(item_list)
        res_item_list = item_list
        item_list = []
        print('=======================')
        # lang = 'en_US'
        prices = get_item_price(res_item_list, lang)
        # print(prices)
        excel = list_to_excel(prices, lang)
        # return JsonResponse({'ok':result.json(),'a':res_item_list})
        return excel
    except Exception as e:
        print(e,'error')
        return JsonResponse({'code':20005,'msg':'导出失败！'})

#获取返回的BOM清单
def get_itemcode(bomlist):
    key = 'childBomList'
    if key in bomlist.keys():
        if bomlist[key]:
            for item in bomlist[key]:
                if item['itemCode']:
                    _item = []
                    _item.append(item['itemCode'])
                    _item.append(item['itemName'])
                    _item.append(item['itemDescription'])
                    valueFlag = 'valueFlag'
                    if valueFlag in item.keys():
                        _item.append(item['valueFlag'])
                    else:
                        _item.append('0')
                    item_list.append(_item)
                    if key in item.keys():
                        get_itemcode(item)

#获取物料的价格
def get_item_price(items,lang):
    prices = []
    tran = {'自研硬件':'Self-developed Hardware','自研软件':'Self-developed Software','外购硬件':'Outsourcing Hardware',
            '外购软件':'Outsourcing Software','定制软件':'Customized Software'}
    db = pymysql.connect(host="192.168.29.73", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    for i in items:
        print(i)
        try:
            sql = 'select productcode,product_type,RRP_USD,RRP_EUR,RRP_GBP,RRP_RMB from myapp_basedata where productcode="%s"' %(str(i[0]))
            cur.execute(sql)
            all_list = list(cur.fetchall())
            print(all_list)
            item = []
            item.append(i[0])
            item.append(i[1])
            item.append(i[2])
            if len(all_list)>0:
                if lang=='en_US':
                    item.append(tran[all_list[0][1].strip()])
                    item.append(all_list[0][2])
                    item.append(all_list[0][3])
                    item.append(all_list[0][4])
                elif lang == 'zh_CN':
                    item.append(all_list[0][1])
                    item.append(all_list[0][5])
                    item.append('')
                    item.append('')
            else:
                _i = ['','','','']
                item = item + _i
            item.append(i[3])
            prices.append(item)
            print('aa')
        except Exception as e:
            print(e,'error')
    return prices

#列表转excel
def list_to_excel(list, lang):
    try:
        wb = openpyxl.Workbook()
        sheel_name = 'sheet1'
        sheet1 = wb.active
        sheet1.title = sheel_name
        sheet1.column_dimensions['A'].width = 18.0
        sheet1.column_dimensions['B'].width = 25.0
        sheet1.column_dimensions['C'].width = 50.0
        sheet1.column_dimensions['D'].width = 20.0
        header1 = ['物料编码', '物料名称', '物料描述', '物料类型', 'RRP_CNY']
        header2 = ['productcode', 'name', 'description', 'product_type', 'RRP_USD', 'RRP_EUR', 'RRP_GBP']
        header3 = ['productcode','name','model','description','product_type','price_source','RRP_USD','RRP_EUR','RRP_GBP','RRP_RMB','createddate','lastmodifieddate','remark']
        ft0 = PatternFill("solid", fgColor="8ac6d1")
        ft2 = Font(name='Calibri', size=11 ,bold=False)
        ft3 = PatternFill("solid", fgColor="9ACD32")
        ft4 = Font(name='Calibri', size=11, bold=True)
        border0 = Border(left=Side(border_style='thin',color='000000'),
                        right=Side(border_style='thin',color='000000'),
                        top=Side(border_style='thin',color='000000'),
                        bottom=Side(border_style='thin',color='000000'))
        border1 = Border(left=Side(border_style='medium', color='000000'),
                         right=Side(border_style='medium', color='000000'),
                         top=Side(border_style='medium', color='000000'),
                         bottom=Side(border_style='medium', color='000000'))
        alignment = Alignment(wrap_text=True)
        if lang == 'zh_CN':
            for i in range(1, len(header1)+1):
                sheet1.cell(row=1, column=i, value=header1[i - 1]).fill = ft0
                sheet1.cell(row=1, column=i).font = ft2
                sheet1.cell(row=1, column=i).border = border1
            for i in range(len(list)):
                if len(list[i])==8 and list[i][7]=='1':
                    for j in range(len(list[i])-3):
                        sheet1.cell(row=i + 2, column=j + 1, value=list[i][j]).font = ft2
                        sheet1.cell(row=i + 2, column=j + 1).border = border0
                        sheet1.cell(row=i + 2, column=j + 1).alignment = alignment
                else:
                    for j in range(len(list[i])-3):
                        sheet1.cell(row=i + 2, column=j + 1, value=list[i][j]).fill = ft3
                        sheet1.cell(row=i + 2, column=j + 1).font = ft4
                        sheet1.cell(row=i + 2, column=j + 1).border = border0
                        sheet1.cell(row=i + 2, column=j + 1).alignment = alignment
        elif lang == 'en_US':
            for i in range(1, len(header2) + 1):
                sheet1.cell(row=1, column=i, value=header2[i - 1]).fill = ft0
                sheet1.cell(row=1, column=i).font = ft2
                sheet1.cell(row=1, column=i).border = border1
            for i in range(len(list)):
                if len(list[i]) == 8 and list[i][7] == '1':
                    for j in range(len(list[i])-1):
                        sheet1.cell(row=i + 2, column=j + 1, value=list[i][j]).font = ft2
                        sheet1.cell(row=i + 2, column=j + 1).border = border0
                        sheet1.cell(row=i + 2, column=j + 1).alignment = alignment
                else:
                    for j in range(len(list[i])-1):
                        sheet1.cell(row=i + 2, column=j + 1, value=list[i][j]).fill = ft3
                        sheet1.cell(row=i + 2, column=j + 1).font = ft4
                        sheet1.cell(row=i + 2, column=j + 1).border = border0
                        sheet1.cell(row=i + 2, column=j + 1).alignment = alignment
        elif lang == 'all':
            for i in range(1, len(header3) + 1):
                sheet1.cell(row=1, column=i, value=header3[i - 1]).fill = ft0
                sheet1.cell(row=1, column=i).font = ft2
            for i in range(len(list)):
                for j in range(len(list[i])):
                    sheet1.cell(row=i + 2, column=j + 1, value=list[i][j]).font = ft2
        excel_name = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d%H%M%S')
        if lang == 'zh_CN':
            excel_name = list[0][1] + u'_价格清单_' + excel_name + '.xlsx'
        elif lang == 'en_US':
            excel_name = list[0][1] + '_price list_' + excel_name + '.xlsx'
        wb.save('test.xlsx')
        file = open('test.xlsx', 'rb')
        response = FileResponse(file)
        print(excel_name)
        response['Content-Type'] = 'application/octet-stream;charset=utf-8'
        response['Content-Disposition'] = "attachment; filename*={}".format(escape_uri_path(excel_name))
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response
    except Exception as e:
        print(e)
        return JsonResponse({'code':20005,'msg':'导出失败！'})

@token_decorator
def create_item(request):
    dict = {}
    dict['productcode'] = request.POST.get('productcode')
    dict['name'] = request.POST.get('name')
    dict['model'] = request.POST.get('model')
    dict['description'] = request.POST.get('description')
    dict['product_type'] = request.POST.get('product_type')
    dict['price_source'] = request.POST.get('price_source')
    dict['RRP_USD'] = request.POST.get('RRP_USD')
    dict['RRP_EUR'] = request.POST.get('RRP_EUR')
    dict['RRP_GBP'] = request.POST.get('RRP_GBP')
    dict['RRP_RMB'] = request.POST.get('RRP_RMB')
    dict['lastmodifieddate'] = datetime.datetime.now()
    dict['remark'] = request.POST.get('remark')
    print(dict['RRP_USD'])
    ret = create_basedata(dict)
    return JsonResponse(ret)

@token_decorator
def update_item(request):
    # print(resolve(request.path_info).url_name)
    try:
        dict = {}
        dict['productcode'] = request.POST.get('productcode')
        dict['name'] = request.POST.get('name')
        dict['model'] = request.POST.get('model')
        dict['description'] = request.POST.get('description')
        dict['product_type'] = request.POST.get('product_type')
        dict['price_source'] = request.POST.get('price_source')
        dict['RRP_USD'] = request.POST.get('RRP_USD')
        dict['RRP_EUR'] = request.POST.get('RRP_EUR')
        dict['RRP_GBP'] = request.POST.get('RRP_GBP')
        dict['RRP_RMB'] = request.POST.get('RRP_RMB')
        dict['lastmodifieddate'] = datetime.datetime.now()
        dict['remark'] = request.POST.get('remark')
        print(dict['productcode'])
        ret = change_basedata(dict)
        return JsonResponse(ret)
    except Exception as e:
        print(e)
        print(traceback.print_exc())
        return JsonResponse({'code': 20006, 'msg': '更新失败！'})

@token_decorator
def delete_item(request):
    id = request.POST.get('id')
    ret = {'code': 20000, 'msg': '删除成功'}
    try:
        if id != None:
            exist_productcode = basedata.objects.filter(id=id)
            if exist_productcode:
                exist = exist_productcode.first()
                dict={}
                dict['productcode'] = exist.productcode
                dict['name'] = exist.name
                dict['model'] = exist.model
                dict['description'] = exist.description
                dict['product_type'] = exist.product_type
                dict['price_source'] = exist.price_source
                dict['RRP_USD'] = ''
                dict['RRP_EUR'] = ''
                dict['RRP_GBP'] = ''
                dict['RRP_RMB'] = ''
                dict=[dict]
                print(dict)
                # update_cpq_item(dict)
                exist_productcode.delete()
                # upsert_cpq_items(dict)
            else:
                ret['code'] = 20005
                ret['msg'] = '删除失败'
        else:
            ret['code'] = 20005
            ret['msg'] = '删除失败'
    except NameError as e:
        print(e)
        ret['code'] = 20006
        ret['msg'] = '未知异常'
    return JsonResponse(ret)

def upsert_basedata(dict):
    sqllist = []
    try:
        if len(dict.get('productcode','')) > 0:
            exist_productcode = basedata.objects.filter(productcode=dict['productcode'])
            if exist_productcode:
                exist_productcode.update(name=dict.get('name',''), model=dict.get('model',''), description=dict.get('description',''),
                                         product_type=dict.get('product_type',''), price_source=dict.get('price_source',''), RRP_USD=dict.get('RRP_USD',''),
                                         RRP_EUR=dict.get('RRP_EUR',''), RRP_GBP=dict.get('RRP_GBP',''), RRP_RMB=dict.get('RRP_RMB',''), remark=dict.get('remark',''),
                                         lastmodifieddate=dict.get('lastmodifieddate',''))
            else:
                sql = basedata(productcode=dict.get('productcode',''),name=dict.get('name',''), model=dict.get('model',''),
                               description=dict.get('description',''), product_type=dict.get('product_type',''), price_source=dict.get('price_source',''),
                               RRP_USD=dict.get('RRP_USD',''), RRP_EUR=dict.get('RRP_EUR',''), RRP_GBP=dict.get('RRP_GBP',''), RRP_RMB=dict.get('RRP_RMB',''),
                               remark=dict.get('remark',''))
                sqllist.append(sql)
        elif len(dict.get('productcode',''))==0 and len(dict.get('model',''))>0:
            exist_model = basedata.objects.filter(Q(productcode=None, model=dict['model'])|Q(productcode='', model=dict['model']))
            if exist_model:
                exist_model.update(productcode=dict.get('productcode',''), name=dict.get('name',''), description=dict.get('description',''),
                                   product_type=dict.get('product_type',''), price_source=dict.get('price_source',''), RRP_USD=dict.get('RRP_USD',''),
                                   RRP_EUR=dict.get('RRP_EUR',''), RRP_GBP=dict.get('RRP_GBP',''), RRP_RMB=dict.get('RRP_RMB',''), remark=dict.get('remark',''),
                                   lastmodifieddate=dict.get('lastmodifieddate',''))
            else:
                sql = basedata(productcode=dict.get('productcode',''), name=dict.get('name',''), model=dict.get('model',''),
                               description=dict.get('description',''), product_type=dict.get('product_type',''), price_source=dict.get('price_source',''),
                               RRP_USD=dict.get('RRP_USD',''), RRP_EUR=dict.get('RRP_EUR',''), RRP_GBP=dict.get('RRP_GBP',''), RRP_RMB=dict.get('RRP_RMB',''),
                               remark=dict.get('remark',''))
                sqllist.append(sql)
    except NameError as e:
        print(e)
    return sqllist

def change_basedata(dict):
    ret = {'code': 20000, 'msg': '修改成功'}
    try:
        if len(dict['productcode']) > 0:
            exist_productcode = basedata.objects.get(productcode=dict['productcode'])
            if exist_productcode:
                # print(exist_productcode.values('id'))
                exist_productcode.__dict__.update(name=dict['name'], model=dict['model'], description=dict['description'],
                                         product_type=dict['product_type'], price_source=dict['price_source'],
                                         RRP_USD=dict['RRP_USD'],
                                         RRP_EUR=dict['RRP_EUR'], RRP_GBP=dict['RRP_GBP'], RRP_RMB=dict['RRP_RMB'],
                                         remark=dict['remark'],
                                         lastmodifieddate=dict['lastmodifieddate'])
                # for i in range(len(exist_productcode)):
                #     exist_productcode[i].name = dict['name']
                #     exist_productcode[i].remark = dict['remark']
                exist_productcode.save()
                # ret['msg'] = update_cpq_item(dict)
                print(ret)
                return ret
            else:
                ret['code'] = 20005
                ret['msg'] = '修改失败'
                return ret
        elif len(dict['productcode'])==0 and len(dict['model'])>0:
            exist_model = basedata.objects.get(Q(productcode=None, model=dict['model'])|Q(productcode='', model=dict['model']))
            if exist_model:
                exist_model.__dict__.update(productcode=dict['productcode'], name=dict['name'], description=dict['description'],
                                   product_type=dict['product_type'], price_source=dict['price_source'],
                                   RRP_USD=dict['RRP_USD'],
                                   RRP_EUR=dict['RRP_EUR'], RRP_GBP=dict['RRP_GBP'], RRP_RMB=dict['RRP_RMB'],
                                   remark=dict['remark'],
                                   lastmodifieddate=dict['lastmodifieddate'])
                # ret['msg'] = update_cpq_item(dict)
                exist_model.save()
                print(ret)
                return ret
            else:
                ret['code'] = 20005
                ret['msg'] = '修改失败'
                return ret
    except NameError as e:
        print(e)
        ret['code'] = 20006
        ret['msg'] = '未知异常'
        return ret

def create_basedata(dict, status=True):
    ret = {'code': 20000, 'msg': '创建成功'}
    try:
        if len(dict['productcode'])>0:
            exist_productcode = basedata.objects.filter(productcode=dict['productcode'])
            if exist_productcode:
                ret['code'] = 20005
                ret['msg'] = '该编码已存在'
                return ret
            else:
                basedata.objects.create(productcode=dict['productcode'], name=dict['name'], model=dict['model'],
                               description=dict['description'],
                               product_type=dict['product_type'], price_source=dict['price_source'],
                               RRP_USD=dict['RRP_USD'],
                               RRP_EUR=dict['RRP_EUR'], RRP_GBP=dict['RRP_GBP'], RRP_RMB=dict['RRP_RMB'],
                               remark=dict['remark'])
                # 如果是CPQ推送过来的价格数据就不创建
                # if status==True:
                #     ret['msg'] = create_cpq_item(dict)
                return ret
        elif len(dict['productcode'])==0 and len(dict['model'])>0:
            exist_model = basedata.objects.filter(Q(productcode=None, model=dict['model'])|Q(productcode='', model=dict['model']))
            if exist_model:
                ret['code'] = 20005
                ret['msg'] = '该型号已存在'
                return ret
            else:
                basedata.objects.create(productcode=dict['productcode'], name=dict['name'], model=dict['model'],
                                        description=dict['description'],
                                        product_type=dict['product_type'], price_source=dict['price_source'],
                                        RRP_USD=dict['RRP_USD'],
                                        RRP_EUR=dict['RRP_EUR'], RRP_GBP=dict['RRP_GBP'], RRP_RMB=dict['RRP_RMB'],
                                        remark=dict['remark'])
                # 如果是CPQ推送过来的价格数据就不创建
                # if status == True:
                #     ret['msg'] = create_cpq_item(dict)
                return ret
    except NameError as e:
        print(e)
        ret['code'] = 20006
        ret['msg'] = '未知异常'
        return ret


#与CPQ的接口
@token_decorator
def get_cpq_token(request):
    access_token = get_cpqq_token()
    return JsonResponse({'access_token':access_token})

def get_cpqq_token():
    url = 'https://xxxxxxxxxxxxxx.com/oauth/oauth/token'
    data = {
        'grant_type':'password',
        'client_id':'hvpc-front-prod',
        'client_secret':'secret',
        'username':'30281805',
        'password':'cHJpY2UxMjM0NTY='
    }
    urllib3.disable_warnings()
    res = requests.post(url, data=data, verify=False)
    return res.json()['access_token']

def get_uat_cpqq_token():
    url = 'https://xxxxxxxxxxxxxx.com/oauth/oauth/token'
    data = {
        'grant_type':'password',
        'client_id':'hvpc-front-uat',
        'client_secret':'secret',
        'username':'30281805',
        'password':'cHJpY2UxMjM0NTY='
    }
    urllib3.disable_warnings()
    res = requests.post(url, data=data, verify=False)
    return res.json()['access_token']

def create_cpq_item(dict,check=False):
    res = {}
    try:
        # price_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price'
        price_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price'
        currency = {'RRP_RMB':'CNY', 'RRP_USD':'USD', 'RRP_EUR':'EUR', 'RRP_GBP':'GBP', 'RRP_AUD':'AUD', 'RRP_CAD':'CAD'}
        ex_currency = {'CNY':'RRP_RMB', 'USD':'RRP_USD', 'EUR':'RRP_EUR', 'GBP':'RRP_GBP', 'AUD':'RRP_AUD', 'CAD':'RRP_CAD'}
        tran = {'自研硬件': 'SelfdevelopedHardware', '自研软件': 'SelfdevelopedSoftware', '外购硬件': 'PurchasedHardware',
                '外购软件': 'PurchasedSoftware', '定制软件': 'CustomizedSoftware','其他硬件':'OtherHardware'}
        #如果是更新接口过来的创建，则不需要再次查询CPQ数据
        if check==False:
            res, headers = get_cpq_item(dict)
        else:
            access_token = get_cpqq_token()
            # access_token = get_uat_cpqq_token()
            headers = {
                'Content-Type': 'application/json; charset=utf-8',
                'Authorization': 'bearer ' + access_token
            }
        # CPQ有价格就不新增
        if len(res.get('content','')) > 0:
            for i in res['content']:
                del dict[ex_currency[i['currencyCode']]]
        n_data = []
        for cur in dict.keys():
            if cur in currency.keys():
                if len(dict.get(cur)) == 0:
                    continue
                else:
                    data = {
                        "enable":"1",
                        "priceCategoryCode":"A",
                        "unitCode":"A",
                        "werksCode":"Hytera",
                        "countryCode":'',
                        "platformCode":dict.get('productcode'),
                        "itemCode": dict.get('productcode'),
                        "itemCharacter": tran[dict.get('product_type')],
                        "model": dict.get('model'),
                        "price": dict.get(cur),
                        "currencyCode":currency[cur],
                        # 'effectStartDate': time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                    }
                    n_data.append(data)
        if len(n_data)>0:
            urllib3.disable_warnings()
            res = requests.post(url=price_url, data=json.dumps(n_data), headers=headers, verify=False)
            print(res.json(),'aaaaaaaaaaaaaaaa')
    except Exception as e:
        print('创建CPQ编码价格异常',e)
        print(traceback.print_exc())
    print(type(res))
    if (isinstance(res,str)) == False:
        if type(res.json()).__name__ == 'dict':
            if 'message' in res.json().keys():
                return res.json()['message']
            else:
                res = '创建至CPQ成功'
                return res
        elif type(res.json()).__name__ == 'list':
            if 'message' in res.json()[0].keys():
                return res.json()[0]['message']
            else:
                res = '创建至CPQ成功'
                return res
    else:
        print('创建',res,type(res))
        return res
def update_cpq_item(dict):
    # price_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price'
    price_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price'
    res = {}
    currency = {'CNY':'RRP_RMB','USD':'RRP_USD','EUR':'RRP_EUR','GBP':'RRP_GBP','AUD':'RRP_AUD','CAD':'RRP_CAD'}
    tran = {'自研硬件': 'SelfdevelopedHardware', '自研软件': 'SelfdevelopedSoftware', '外购硬件': 'PurchasedHardware',
            '外购软件': 'PurchasedSoftware', '定制软件': 'CustomizedSoftware', '其他硬件': 'OtherHardware'}
    try:
        ress,headers = get_cpq_item(dict)
        c_data = []
        exist_currency = []
        #CPQ有价格就更新
        for res in ress:
            if len(res['content']) > 0:
                for i in res['content']:
                    print(i)
                    if i['currencyCode'] in currency.keys():
                        enable = 1
                        if len(dict.get(currency[i['currencyCode']], '')) == 0:
                            enable = 0
                        u_data = {
                            'objectVersionNumber': i['objectVersionNumber'],
                            'id': i['id'],
                            'price': dict.get(currency[i['currencyCode']], ''),
                            'currencyCode': i['currencyCode'],
                            'unitCode': i['unitCode'],
                            'priceCategoryCode': i['priceCategoryCode'],
                            'itemCharacter': tran[dict.get('product_type')],
                            'itemCode': i.get('itemCode'),
                            'model': i.get('model'),
                            'effectStartDate':time.strftime("%Y-%m-%d %H:%M:%S",time.localtime()),
                            'enable':enable
                        }
                        c_data.append(u_data)
                    exist_currency.append(currency[i['currencyCode']])
        #更新价格
        print('aaaa',c_data)
        if len(c_data) > 0:
            urllib3.disable_warnings()
            res = requests.put(url=price_url, data=json.dumps(c_data), headers=headers, verify=False)
            print('更新CPQ价格',res.json())
        #平台有，CPQ没有，则CPQ新增价格
        exist_currency = list(set(exist_currency))
        print(exist_currency)
        if len(exist_currency)>0:
            for j in exist_currency:
                del dict[j]
        for item in currency.values():
            if item not in dict.keys():
                continue
            else:
                # res = create_cpq_item(dict,True)
                break
    except Exception as e:
        print('更新CPQ编码价格异常',e)
        print(traceback.print_exc())
        res = str(e)
    print(type(res),'wwwwwwwwwww')
    if (isinstance(res,str)) == False:
        if type(res.json()).__name__ == 'dict':
            if 'message' in res.json().keys():
                return res.json()['message']
            else:
                res = '更新至CPQ成功'
                return res
        elif isinstance(res.json(),list):
            if 'message' in res.json()[0].keys():
                return res.json()[0]['message']
            else:
                res = '更新至CPQ成功'
                return res
        print('back',type(res.json()))
    else:
        print('更新',res,type(res))
        return res
def get_cpq_item(dict):
    access_token = get_cpqq_token()
    # access_token = get_uat_cpqq_token()
    headers = {
        'Content-Type': 'application/json; charset=utf-8',
        'Authorization': 'bearer ' + access_token
    }
    search_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price/list'
    # search_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price/list'
    urllib3.disable_warnings()
    res={}
    ress = []
    try:
        s_data = {}
        if len(dict.get('productcode','')) > 0:
            itemCode = dict['productcode']
            s_data = {'itemCode':itemCode, 'page':0, 'showAll':1, 'size':10}
        elif len(dict.get('productcode'))==0 and len(dict.get('model',''))>0:
            model = dict['model']
            s_data = {'model':model, 'page':0, 'showAll':1, 'size':10}
        response = requests.get(url=search_url, params=s_data, headers=headers, verify=False)
        res = response.json()
        ress.append(res)
        if len(dict.get('productcode'))>0 and len(dict.get('model',''))>0:
            model = dict['model']
            s_data = {'model': model, 'page': 0, 'showAll': 1, 'size': 10}
            response = requests.get(url=search_url, params=s_data, headers=headers, verify=False)
            res = response.json()
            ress.append(res)
        print(ress)
    except Exception as e:
        print(e)
    finally:
        return ress,headers

def upsert_cpq_items(dict):
    try:
        print('要新增到CPQ的数据：',dict)
        # access_token = get_cpqq_token()
        # access_token = get_uat_cpqq_token()
        access_token = '1'
        headers = {
            'Content-Type': 'application/json; charset=utf-8',
            'Authorization': 'bearer ' + access_token
        }
        # upsert_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price/update-batch'
        upsert_url = 'http://10.160.26.56:9020/v0/11/3/quote/mdm/item/price/update-batch'
        # upsert_url = 'https://xxxxxxxxxxxxxx.com/hvpc-quote/v1/11/3/quote/mdm/item/price/update-batch'
        urllib3.disable_warnings()
        currency = {'RRP_RMB':'CNY', 'RRP_USD':'USD', 'RRP_EUR':'EUR', 'RRP_GBP':'GBP', 'RRP_AUD':'AUD', 'RRP_CAD':'CAD'}
        tran = {'自研硬件': 'SelfdevelopedHardware', '自研软件': 'SelfdevelopedSoftware', '外购硬件': 'PurchasedHardware',
                '外购软件': 'PurchasedSoftware', '定制软件': 'CustomizedSoftware', '其他硬件': 'OtherHardware'}
        n_data = []
        all_cur = currency.keys()
        for item in dict:
            for cur in item.keys():
                if cur in all_cur:
                    enable = 1
                    model = item.get('model','')
                    price = item.get(cur, '').strip()
                    if len(price) == 0 or price == '/':
                        enable = 0
                        price = None
                    if len(item.get('productcode', '')) > 0:
                        model = ''
                    if model=='' and len(item.get('productcode', ''))==0:
                        continue
                    u_data = {
                        "werksCode":"Hytera",
                        "price": price,
                        "currencyCode": currency[cur],
                        "itemCharacter": tran.get(item.get('product_type','').strip(),''),
                        "itemCode": item.get('productcode',''),
                        "model": model,
                        'unitCode': "A",
                        "enable": enable
                    }
                    n_data.append(u_data)
        print('insert data is: ',n_data)
        if len(n_data) > 0:
            res = requests.put(url=upsert_url, data=json.dumps(n_data), headers=headers, verify=False)
            print(res.json(), 'aaaaaaaaaaaaaaaa')
            return res.json()
    except Exception as e:
        print('批量更新异常:',e)
        print(traceback.print_exc())
        return '批量更新异常'

#CPQ推送创建价格数据
@token_decorator
def create_price_item(request):
    ret = {}
    try:
        dict = {}
        dict['productcode'] = request.POST.get('itemCode','')
        dict['name'] = request.POST.get('itemName','')
        dict['model'] = request.POST.get('model','')
        dict['description'] = request.POST.get('description','')
        dict['product_type'] = '自研硬件'
        dict['price_source'] = 'CPQ推送'
        dict['RRP_USD'] = request.POST.get('USD','')
        dict['RRP_EUR'] = request.POST.get('EUR','')
        dict['RRP_GBP'] = request.POST.get('GBP','')
        dict['RRP_RMB'] = request.POST.get('CNY','')
        dict['lastmodifieddate'] = datetime.datetime.now()
        dict['remark'] = request.POST.get('remark','')
        ret = create_basedata(dict,False)
    except Exception as e:
        print(e)
        ret = {'code':20006,'msg':'推送价格失败！'}
    return JsonResponse(ret)


