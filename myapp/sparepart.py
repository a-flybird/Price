# -*- coding: utf-8 -*-
# @Author      ：180732677
# @Email       : yinbao.bai@hytera.com
# @Time        ：2020/6/19  16:54
# @ProjectName : mysite
# @FileName    ：sparepart.py
# @Software    ：PyCharm
from django.core.paginator import Paginator
from django.http import HttpResponse,JsonResponse,FileResponse
from django.utils.encoding import escape_uri_path
from lxml.doctestcompare import strip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill,Font

from .models import basedata,BJ_basedata
from openpyxl.compat import range
from .views import token_decorator
import requests,urllib3,openpyxl,pymysql,csv,datetime,codecs

def download_user(request):
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    sql = 'select productcode,description,description_en,unit,USD,EUR,GBP,CNY,CAD,AUD,createddate,lastmodifieddate from myapp_bj_basedata '
    cur.execute(sql)
    data = list(cur.fetchall())
    response = list_to_excel(data,lang='all')
    return response

@token_decorator
def upsert_user(request):
    uploadedFile = request.FILES.get('file', None)
    print(uploadedFile)
    wb = load_workbook(filename=uploadedFile)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    rows = ws.max_row
    cols = ws.max_column
    headers = ['productcode','description','description_en','unit','USD','EUR','GBP','CNY','CAD','AUD']
    lists = []
    if rows-1 == 0 and cols-1 == 0:
        pass
    else:
        sqllist = []
        for i in range(2, rows + 1):
            r = {}
            for j in range(1, len(headers) + 1):
                key = headers[j - 1]
                if ws.cell(i, j):
                    r[key] = ws.cell(i, j).value
                else:
                    r[key] = ' '
            lists.append(r)
        db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
        cur = db.cursor()
        for cell in lists:
            cell['lastmodifieddate'] = datetime.datetime.now()
            sqllist = sqllist + upsert_basedata(cell)
        if len(sqllist) > 0:
            BJ_basedata.objects.bulk_create(sqllist)
    return JsonResponse({'ok':'ok'})

@token_decorator
def price_list(request):
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    sql = 'select productcode,description,description_en,unit,USD,EUR,GBP,CNY,CAD,AUD,createddate,lastmodifieddate,id from myapp_bj_basedata'
    cur.execute(sql)
    data = cur.fetchall()
    total = len(data)
    page_size = request.GET.get('page_size',10)
    paginator = Paginator(data,page_size)
    page_num = request.GET.get('page_num',1)
    context = []
    try:
        if paginator.page(page_num):
            page_of_price = paginator.page(page_num)
            context=list(page_of_price)
    except Exception as e:
        print(e)
        page_num = 1
        page_of_price = paginator.page(page_num)
        context = list(page_of_price)
    # print(context)
    jsonData = []
    for row in context:
        res = {}
        res['productcode'] = row[0]
        res['description'] = row[1]
        res['description_en'] = row[2]
        res['unit'] = row[3]
        res['USD'] = row[4]
        res['EUR'] = row[5]
        res['GBP'] = row[6]
        res['CNY'] = row[7]
        res['CAD'] = row[8]
        res['AUD'] = row[9]
        res['createddate'] = row[10]
        res['lastmodifieddate'] = row[11]
        res['id'] = row[12]
        jsonData.append(res)
    # print(jsonData)
    return JsonResponse({'rows':jsonData,'page_num':int(page_num),'total':total})

@token_decorator
def create_item(request):
    try:
        dict = {}
        dict['productcode'] = request.POST.get('productcode')
        dict['description'] = request.POST.get('description')
        dict['description_en'] = request.POST.get('description_en')
        dict['unit'] = request.POST.get('unit')
        dict['USD'] = request.POST.get('USD')
        dict['EUR'] = request.POST.get('EUR')
        dict['GBP'] = request.POST.get('GBP')
        dict['CNY'] = request.POST.get('CNY')
        dict['CAD'] = request.POST.get('CAD')
        dict['AUD'] = request.POST.get('AUD')
        ret = create_basedata(dict)
        return JsonResponse(ret)
    except Exception as e:
        print(e,'123')
        return JsonResponse({'code': 20006, 'msg': '未知异常'})

@token_decorator
def update_item(request):
    dict = {}
    dict['productcode'] = request.POST.get('productcode')
    dict['description'] = request.POST.get('description')
    dict['description_en'] = request.POST.get('description_en')
    dict['unit'] = request.POST.get('unit')
    dict['USD'] = request.POST.get('USD')
    dict['EUR'] = request.POST.get('EUR')
    dict['GBP'] = request.POST.get('GBP')
    dict['CNY'] = request.POST.get('CNY')
    dict['CAD'] = request.POST.get('CAD')
    dict['AUD'] = request.POST.get('AUD')
    dict['lastmodifieddate'] = datetime.datetime.now()
    ret = change_basedata(dict)
    return JsonResponse(ret)

@token_decorator
def delete_item(request):
    id = request.POST.get('id')
    ret = {'code': 20000, 'msg': '删除成功'}
    try:
        if id != None:
            exist_productcode = BJ_basedata.objects.filter(id=id)
            if exist_productcode:
                exist_productcode.delete()
            else:
                ret['code'] = 20005
                ret['msg'] = '删除失败'
        else:
            ret['code'] = 20005
            ret['msg'] = '删除失败'
    except NameError as e:
        print(e)
        ret['code'] = 20006
        ret['msg'] = '未知异常'
    return JsonResponse(ret)

def upsert_basedata(dict):
    sqllist = []
    try:
        if dict['productcode'] != None:
            exist_productcode = BJ_basedata.objects.filter(productcode=dict['productcode'])
            if exist_productcode:
                exist_productcode.update(description=dict['description'], description_en=dict['description_en'],
                               unit=dict['unit'],USD=dict['USD'], EUR=dict['EUR'],GBP=dict['GBP'],
                               CNY=dict['CNY'], CAD=dict['CAD'], AUD=dict['AUD'],lastmodifieddate=dict['lastmodifieddate'])
            else:
                sql = BJ_basedata(productcode=dict['productcode'], description=dict['description'], description_en=dict['description_en'],
                               unit=dict['unit'],USD=dict['USD'], EUR=dict['EUR'],GBP=dict['GBP'],
                               CNY=dict['CNY'], CAD=dict['CAD'], AUD=dict['AUD'])
                sqllist.append(sql)
    except NameError as e:
        print(e)
    return sqllist

def change_basedata(dict):
    ret = {'code': 20000, 'msg': '修改成功'}
    try:
        if dict['productcode'] != None:
            exist_productcode = BJ_basedata.objects.filter(productcode=dict['productcode'])
            if exist_productcode:
                exist_productcode.update(description=dict['description'], description_en=dict['description_en'],
                               unit=dict['unit'],USD=dict['USD'], EUR=dict['EUR'],GBP=dict['GBP'],
                               CNY=dict['CNY'], CAD=dict['CAD'], AUD=dict['AUD'],lastmodifieddate=dict['lastmodifieddate'])
                return ret
            else:
                ret['code'] = 20005
                ret['msg'] = '修改失败'
                return ret
    except NameError as e:
        print(e)
        ret['code'] = 20006
        ret['msg'] = '未知异常'
        return ret

def create_basedata(dict):
    ret = {'code': 20000, 'msg': '创建成功'}
    try:

        if dict['productcode'] != None:
            exist_productcode = BJ_basedata.objects.filter(productcode=dict['productcode'])
            if exist_productcode:
                ret['code'] = 20005
                ret['msg'] = '该编码已存在'
                return ret
            else:
                BJ_basedata.objects.create(productcode=dict['productcode'], description=dict['description'], description_en=dict['description_en'],
                               unit=dict['unit'],USD=dict['USD'], EUR=dict['EUR'],GBP=dict['GBP'],
                               CNY=dict['CNY'], CAD=dict['CAD'], AUD=dict['AUD'])
                return ret
    except NameError as e:
        print(e,'ERROR')
        ret['code'] = 20006
        ret['msg'] = '未知异常'
        return ret

#列表转excel
def list_to_excel(list, lang):
    try:
        wb = openpyxl.Workbook()
        sheel_name = 'sheet1'
        sheet1 = wb.active
        sheet1.title = sheel_name
        sheet1.column_dimensions['A'].width = 18.0
        sheet1.column_dimensions['B'].width = 50.0
        sheet1.column_dimensions['C'].width = 50.0
        header3 = ['productcode','description','description_en','unit','USD','EUR','GBP','CNY','CAD','AUD','createddate','lastmodifieddate']
        ft0 = PatternFill("solid", fgColor="8ac6d1")
        ft2 = Font(name='Calibri', size=11 ,bold=False)
        if lang == 'all':
            for i in range(1, len(header3) + 1):
                sheet1.cell(row=1, column=i, value=header3[i - 1]).fill = ft0
                sheet1.cell(row=1, column=i).font = ft2
            for i in range(len(list)):
                for j in range(len(list[i])):
                    sheet1.cell(row=i + 2, column=j + 1, value=list[i][j]).font = ft2
        excel_name = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d%H%M%S')
        wb.save('test.xlsx')
        file = open('test.xlsx', 'rb')
        response = FileResponse(file)
        print(excel_name)
        response['Content-Type'] = 'application/octet-stream;charset=utf-8'
        response['Content-Disposition'] = "attachment; filename*={}".format(escape_uri_path(excel_name))
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response
    except Exception as e:
        print(e)
        return JsonResponse({'code':20005,'msg':'导出失败！'})



@token_decorator
def search(request):
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()
    results = []
    jsonData = []
    if request.method == "GET":
        start = request.GET.get("productcode", None)
        if start is not None and len(str(start))!=0:
            starts = start.split(',')
            for i in starts:
                starts = strip(i)
                sql = "SELECT id,productcode,description,description_en,unit,USD,EUR,GBP,CNY,CAD,AUD FROM myapp_bj_basedata WHERE productcode like '%" \
                      + starts + "%' OR description like '%" + starts + "%'"
                try:
                    cur.execute(sql)
                    results = results + list(cur.fetchall())
                    print(results)
                except Exception as e:
                    raise e
            for row in results:
                res = {}
                res['id'] = row[0]
                res['productcode'] = row[1]
                res['description'] = row[2]
                res['description_en'] = row[3]
                res['unit'] = row[4]
                res['USD'] = row[5]
                res['EUR'] = row[6]
                res['GBP'] = row[7]
                res['CNY'] = row[8]
                res['CAD'] = row[9]
                res['AUD'] = row[10]
                jsonData.append(res)
            print(jsonData)
        else:
            print('未找到记录')
    return JsonResponse({'rows':jsonData})
@token_decorator
def exportExcel(request):
    file = open('E:/mysite/test2.xlsx', 'rb')
    response = HttpResponse(file)
    response['Content-Type'] = 'application/octet-stream'
    response['Content-Disposition'] = 'attachment;filename="File template.xlsx"'
    return response
@token_decorator
def uploadExcel(request):
    print('123123132')
    uploadedFile = request.FILES.get('file', None)
    print(uploadedFile)
    wb = openpyxl.load_workbook(filename=uploadedFile)
    ws = wb.get_sheet_names()
    ws = wb.get_sheet_by_name(ws[0])
    headers = ['productcode','description','description_en','unit','USD','EUR','GBP','CNY','CAD','AUD']
    lists = []
    rows = ws.max_row
    for row in range(2, rows + 1):
        r = {}
        for col in range(1, len(headers) + 1):
            key = headers[col - 1]
            r[key] = ws.cell(row=row, column=col).value
        lists.append(r)
    db = pymysql.connect(host="127.0.0.1", user="root", password="******************", db="MYDATA", port=3306)
    cur = db.cursor()

    wb = openpyxl.Workbook()
    sheel_name = 'test'
    sheet1 = wb.active
    sheet1.title = sheel_name
    row = 2
    ft0 = openpyxl.styles.PatternFill("solid", fgColor="8ac6d1")
    ft1 = openpyxl.styles.PatternFill("solid", fgColor="ffb6b9")
    ft3 = openpyxl.styles.PatternFill("solid", fgColor="fdcb6e")
    ft2 = openpyxl.styles.Font(name='Calibri', size=11)
    # 写入表头
    for i in range(1, 11):
        sheet1.cell(row=1, column=i, value=headers[i - 1]).fill = ft0
    for cell in lists:
        productcode = cell['productcode']
        description = cell['description']
        description_en = cell['description_en']
        unit = cell['unit']
        USD = cell['USD']
        EUR = cell['EUR']
        GBP = cell['GBP']
        CNY = cell['CNY']
        CAD = cell['CAD']
        AUD = cell['AUD']
        try:
            all_list = BJ_basedata.objects.get(productcode=productcode)
            if all_list:
                sheet1.cell(row=row, column=1, value=all_list.productcode)
                sheet1.cell(row=row, column=2, value=all_list.description)
                sheet1.cell(row=row, column=3, value=all_list.description_en)
                sheet1.cell(row=row, column=4, value=all_list.unit)
                if str(USD) == all_list.USD:
                    sheet1.cell(row=row, column=5, value=USD)
                else:
                    sheet1.cell(row=row, column=5, value=all_list.USD).fill = ft3
                if str(EUR) == all_list.EUR:
                    sheet1.cell(row=row, column=6, value=EUR)
                else:
                    sheet1.cell(row=row, column=6, value=all_list.EUR).fill = ft3
                if str(GBP) == all_list.GBP:
                    sheet1.cell(row=row, column=7, value=GBP)
                else:
                    sheet1.cell(row=row, column=7, value=all_list.GBP).fill = ft3
                if str(CNY) == all_list.CNY:
                    sheet1.cell(row=row, column=8, value=CNY)
                else:
                    sheet1.cell(row=row, column=8, value=all_list.CNY).fill = ft3
                if str(CAD) == all_list.CAD:
                    sheet1.cell(row=row, column=9, value=CAD)
                else:
                    sheet1.cell(row=row, column=9, value=all_list.CAD).fill = ft3
                if str(AUD) == all_list.AUD:
                    sheet1.cell(row=row, column=10, value=AUD)
                else:
                    sheet1.cell(row=row, column=10, value=all_list.AUD).fill = ft3
        except:
            sheet1.cell(row=row, column=1, value=str(productcode)).fill = ft1
            sheet1.cell(row=row, column=2, value=description).fill = ft1
            sheet1.cell(row=row, column=3, value=description_en).fill = ft1
            sheet1.cell(row=row, column=4, value=unit).fill = ft1
            sheet1.cell(row=row, column=5, value=USD).fill = ft1
            sheet1.cell(row=row, column=6, value=EUR).fill = ft1
            sheet1.cell(row=row, column=7, value=GBP).fill = ft1
            sheet1.cell(row=row, column=8, value=CNY).fill = ft1
            sheet1.cell(row=row, column=9, value=CAD).fill = ft1
            sheet1.cell(row=row, column=10, value=AUD).fill = ft1
        row = row + 1
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            sheet1.cell(row, col).font = ft2
    excel_name = datetime.datetime.strftime(datetime.datetime.now(), '%Y%m%d%H%M%S')
    wb.save('test.xlsx')
    file = open('test.xlsx','rb')
    response = FileResponse(file)
    response['Content-Type'] = 'application/octet-stream;charset=UTF-8'
    response['Content-Disposition'] = 'attachment; filename=%s.xlsx' % excel_name
    response['Access-Control-Expose-Headers'] = 'Content-Disposition'
    return response
